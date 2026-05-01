"""
=============================================================
  Certificate Generator - ONSkills Skill Competency Certificate
=============================================================
  Reads candidate data from Excel, overlays text on certificate
  template, and generates PDF certificates.

  Usage:
    python generate_certificates.py                    # Normal mode
    python generate_certificates.py --calibrate        # Shows field positions on template
    python generate_certificates.py --excel data.xlsx  # Use specific Excel file

  Author: Auto-generated
  Date: 2026-05-01
=============================================================
"""

import os
import sys
import glob
import argparse

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook


# =============================================================
#  CONFIGURATION - Adjust these values to fine-tune positions
# =============================================================

# Template image path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "template", "certificate_template.jpg")

# Input/Output folders
INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input")
OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "output")

# Font settings
# Using Arial Bold from Windows fonts
FONT_PATH = r"C:\Windows\Fonts\arialbd.ttf"
FONT_SIZE = 18          # Default font size for most fields
FONT_SIZE_SMALL = 16    # Smaller font for fields with less space
FONT_SIZE_NAME = 20     # Slightly bigger for candidate name
FONT_COLOR = (0, 0, 0)  # Black text

# =============================================================
#  FIELD POSITIONS - (x, y) coordinates on the 1312x945 image
#  Adjust these to perfectly align text on your template
# =============================================================

FIELD_POSITIONS = {
    # Field name         :  (x,    y,    font_size,  max_width)
    "Certificate No"     :  (1090, 199,  FONT_SIZE_SMALL, 150),
    "Name"               :  (680,  331,  FONT_SIZE_NAME,  420),
    "Father/Spouse Name" :  (479,  400,  FONT_SIZE,       440),
    "Aadhar No"          :  (970,  400,  FONT_SIZE,       320),
    "Job Role"           :  (940,  461,  FONT_SIZE_SMALL, 810),
    "Duration"           :  (409,  525,  FONT_SIZE_SMALL, 120),
    "Training Center"    :  (708,  525,  FONT_SIZE_SMALL, 415),
    "Sponsored By"       :  (490,  583,  FONT_SIZE_SMALL, 220),
    "District"           :  (860,  583,  FONT_SIZE_SMALL, 160),
    "State"              :  (1095, 583,  FONT_SIZE_SMALL, 230),
    "Place of Issue"     :  (555,  813,  FONT_SIZE_SMALL, 220),
    "Date of Issue"      :  (560,  856,  FONT_SIZE_SMALL, 220),
}



# Excel column headers (must match exactly with your Excel file)
EXCEL_HEADERS = [
    "Certificate No",
    "Name",
    "Title",              # Mr. / Ms. / Mrs.
    "Father/Spouse Name",
    "Relation",           # S/o, D/o, W/o
    "Aadhar No",
    "Job Role",
    "Duration",
    "Training Center",
    "Sponsored By",
    "District",
    "State",
    "Place of Issue",
    "Date of Issue",
]


# =============================================================
#  HELPER FUNCTIONS
# =============================================================

def get_font(size):
    """Load the font at the given size. Falls back to default if not found."""
    try:
        return ImageFont.truetype(FONT_PATH, size)
    except (IOError, OSError):
        # Try alternate paths
        alternate_fonts = [
            r"C:\Windows\Fonts\arial.ttf",
            r"C:\Windows\Fonts\calibrib.ttf",
            r"C:\Windows\Fonts\segoeui.ttf",
        ]
        for alt_font in alternate_fonts:
            try:
                return ImageFont.truetype(alt_font, size)
            except (IOError, OSError):
                continue
        print("[WARNING] No system font found! Using default bitmap font.")
        return ImageFont.load_default()


def draw_text_fitted(draw, x, y, text, font_size, max_width, color=FONT_COLOR):
    """
    Draw text at (x, y) position. If text is too wide, reduce font size
    to fit within max_width.
    """
    current_size = font_size
    font = get_font(current_size)

    # Measure text width and reduce font size if needed
    while current_size > 10:
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        if text_width <= max_width:
            break
        current_size -= 1
        font = get_font(current_size)

    draw.text((x, y), text, font=font, fill=color)
    return font


def read_excel(filepath):
    """Read candidate data from Excel file. Returns list of dicts."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # Read headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")

    print(f"\n  Excel headers found: {headers}")

    # Read data rows
    candidates = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        data = {}
        is_empty = True
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                if cell.value is not None:
                    if isinstance(cell.value, datetime):
                        # Format: 20 Apr 2026
                        value = cell.value.strftime("%d %b %Y")
                    else:
                        value = str(cell.value).strip()
                else:
                    value = ""
                
                data[headers[col_idx]] = value
                if value:
                    is_empty = False

        if not is_empty:
            data["_row"] = row_idx
            candidates.append(data)

    wb.close()
    return candidates


def generate_certificate(template_img, candidate_data, output_path):
    """Generate a single certificate for one candidate with auto-upscaling."""
    
    # SCALE_FACTOR: 2.0x makes the 1312px width image approx 2624px (Good print quality)
    SCALE_FACTOR = 2.0
    
    # Upscale the template for better physical size and crispness
    new_size = (int(template_img.width * SCALE_FACTOR), int(template_img.height * SCALE_FACTOR))
    img = template_img.resize(new_size, Image.Resampling.LANCZOS)
    draw = ImageDraw.Draw(img)

    # Build display values
    display_data = dict(candidate_data)

    # Combine Title + Name
    title = candidate_data.get("Title", "").strip()
    name = candidate_data.get("Name", "").strip()
    if title:
        display_data["Name"] = f"{title} {name}"

    # Combine Relation + Father/Spouse Name
    relation = candidate_data.get("Relation", "").strip()
    parent = candidate_data.get("Father/Spouse Name", "").strip()
    if relation:
        display_data["Father/Spouse Name"] = f"{relation} {parent}"
    
    # Clean Certificate No if it contains the label
    cert_val = candidate_data.get("Certificate No", "")
    if ":" in cert_val:
        cert_val = cert_val.split(":")[-1].strip()
    display_data["Certificate No"] = cert_val

    # Draw each field with scaled coordinates
    for field_name, (x, y, font_size, max_width) in FIELD_POSITIONS.items():
        text = display_data.get(field_name, "")
        if text and text != "None":
            # Apply scaling to all parameters
            scaled_x = int(x * SCALE_FACTOR)
            scaled_y = int(y * SCALE_FACTOR)
            scaled_font_size = int(font_size * SCALE_FACTOR)
            scaled_max_width = int(max_width * SCALE_FACTOR)
            
            draw_text_fitted(draw, scaled_x, scaled_y, text, scaled_font_size, scaled_max_width)

    # Save as PDF
    if img.mode == "RGBA":
        img = img.convert("RGB")

    img.save(output_path, "PDF", resolution=300.0, subsampling=0, quality=100)
    return True


def calibrate_mode():
    """
    Calibration mode: Draws red markers and labels at all field positions
    so you can visually verify and adjust coordinates.
    """
    print("\n  === CALIBRATION MODE ===")

    template = Image.open(TEMPLATE_PATH)
    draw = ImageDraw.Draw(template)
    font = get_font(12)

    for field_name, (x, y, font_size, max_width) in FIELD_POSITIONS.items():
        # Draw a red crosshair at the position
        cross_size = 8
        draw.line([(x - cross_size, y), (x + cross_size, y)], fill=(255, 0, 0), width=2)
        draw.line([(x, y - cross_size), (x, y + cross_size)], fill=(255, 0, 0), width=2)

        # Draw the max_width boundary
        draw.rectangle([(x, y), (x + max_width, y + font_size + 4)],
                       outline=(255, 0, 0, 128), width=1)

        # Label the field
        draw.text((x, y - 15), f"{field_name}", font=font, fill=(255, 0, 0))

        # Draw sample text
        sample_font = get_font(font_size)
        draw.text((x, y), "SAMPLE TEXT", font=sample_font, fill=(0, 0, 200))

    output_path = os.path.join(OUTPUT_FOLDER, "_calibration_preview.jpg")
    template.save(output_path, quality=95)
    print(f"\n  Calibration image saved to: {output_path}")
    print("  Open it to check if text positions are correct.")
    print("  Adjust FIELD_POSITIONS in the script if needed.\n")


def create_sample_excel():
    """Create a sample Excel file with dummy data for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Write headers
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Sample data
    sample_data = [
        ["SO/CV/2026/01", "Rahul Kumar", "Mr.", "Ram Kumar", "S/o",
         "1234 5678 9012", "Solar Panel Installation Technician",
         "600 Hours", "ABC Training Center, Delhi", "NSDC",
         "South Delhi", "Delhi", "New Delhi", "01/05/2026"],
        ["SO/CV/2026/02", "Priya Sharma", "Ms.", "Rajesh Sharma", "D/o",
         "9876 5432 1098", "Electrician Domestic Solutions",
         "400 Hours", "XYZ Skill Center, Mumbai", "PMKVY",
         "Mumbai", "Maharashtra", "Mumbai", "01/05/2026"],
        ["SO/CV/2026/03", "Suman Devi", "Mrs.", "Anil Singh", "W/o",
         "5555 6666 7777", "Self Employed Tailor",
         "300 Hours", "PQR Training Institute, Jaipur", "DDU-GKY",
         "Jaipur", "Rajasthan", "Jaipur", "01/05/2026"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        max_len = len(header)
        for row_data in sample_data:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    output_path = os.path.join(INPUT_FOLDER, "sample_candidates.xlsx")
    wb.save(output_path)
    print(f"\n  Sample Excel created: {output_path}")
    return output_path


# =============================================================
#  MAIN
# =============================================================

def main():
    parser = argparse.ArgumentParser(description="ONSkills Certificate Generator")
    parser.add_argument("--calibrate", action="store_true",
                        help="Generate calibration preview to check text positions")
    parser.add_argument("--excel", type=str, default=None,
                        help="Path to Excel file (default: auto-detect from input/ folder)")
    parser.add_argument("--sample", action="store_true",
                        help="Create a sample Excel file in input/ folder")
    args = parser.parse_args()

    print("\n  ================================================")
    print("  |  ONSkills Certificate Generator v1.0         |")
    print("  |  Cee Vision Technologies Pvt. Ltd.           |")
    print("  ================================================")

    # Ensure folders exist
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Check template exists
    if not os.path.exists(TEMPLATE_PATH):
        print(f"\n  [ERROR] Template not found: {TEMPLATE_PATH}")
        print("  Please place your certificate template in the 'template/' folder.")
        sys.exit(1)

    # Sample mode
    if args.sample:
        create_sample_excel()
        return

    # Calibration mode
    if args.calibrate:
        calibrate_mode()
        return

    # Find Excel file
    excel_path = args.excel
    if not excel_path:
        # Auto-detect from input folder
        excel_files = glob.glob(os.path.join(INPUT_FOLDER, "*.xlsx"))
        if not excel_files:
            print(f"\n  [ERROR] No Excel file found in: {INPUT_FOLDER}")
            print("  Place your Excel file in the 'input/' folder, or use --excel flag.")
            print("  Use --sample to create a sample Excel file.\n")
            sys.exit(1)
        if len(excel_files) > 1:
            print(f"\n  [WARNING] Multiple Excel files found. Using: {os.path.basename(excel_files[0])}")
        excel_path = excel_files[0]

    if not os.path.exists(excel_path):
        print(f"\n  [ERROR] Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"\n  Template : {TEMPLATE_PATH}")
    print(f"  Excel    : {excel_path}")
    print(f"  Output   : {OUTPUT_FOLDER}")

    # Load template
    print("\n  Loading template...")
    template = Image.open(TEMPLATE_PATH)
    print(f"  Template size: {template.size[0]} x {template.size[1]} px")

    # Read Excel data
    print("  Reading Excel data...")
    candidates = read_excel(excel_path)

    if not candidates:
        print("\n  [ERROR] No candidate data found in Excel!")
        sys.exit(1)

    print(f"  Found {len(candidates)} candidate(s)")

    # Generate certificates
    print("\n  -------------------------------------------")
    print("  Generating Certificates...")
    print("  -------------------------------------------\n")

    success_count = 0
    error_count = 0

    for idx, candidate in enumerate(candidates, start=1):
        name = candidate.get("Name", f"Unknown_{idx}")
        cert_no = candidate.get("Certificate No", f"CERT_{idx:03d}")
        issue_date = candidate.get("Date of Issue", "No_Date").strip()

        # Clean cert_no for filename
        if ":" in cert_no:
            cert_no = cert_no.split(":")[-1].strip()

        # Create subfolder for Date of Issue
        # Clean date string for folder name (remove / or : if any)
        safe_date = issue_date.replace("/", "-").replace(":", "-").replace(" ", "_")
        date_folder = os.path.join(OUTPUT_FOLDER, safe_date)
        os.makedirs(date_folder, exist_ok=True)

        # Clean filename (remove special characters)
        safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_cert = "".join(c for c in cert_no if c.isalnum() or c in ('-', '_', '/')).strip()
        safe_cert = safe_cert.replace("/", "-")

        # Filename and output path inside date folder
        filename = f"{safe_name}_{safe_cert}.pdf"
        output_path = os.path.join(date_folder, filename)

        try:
            generate_certificate(template, candidate, output_path)
            print(f"  [OK] [{idx}/{len(candidates)}] {name} -> {filename}")
            success_count += 1
        except Exception as e:
            print(f"  [FAIL] [{idx}/{len(candidates)}] {name} -> ERROR: {e}")
            error_count += 1

    # Summary
    print("\n  ================================================")
    print(f"  DONE! Generated: {success_count} | Errors: {error_count}")
    print(f"  Output folder: {OUTPUT_FOLDER}")
    print("  ================================================\n")


if __name__ == "__main__":
    main()
